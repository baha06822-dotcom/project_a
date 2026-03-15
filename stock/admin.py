from django.contrib import admin
from django.urls import path
from django.shortcuts import redirect
from django.utils.html import format_html
from django.http import HttpResponse, JsonResponse
from django.utils.http import urlencode

import json
import copy

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import CurrentStock
from .admin_import import register_stock_import_page


# -----------------------------
# helpers
# -----------------------------
def _clean_single_int(value: str | None) -> str | None:
    """
    Нормализует значение id из querystring.
    Исправляет случаи вида "['1']" -> "1".
    Возвращает строку с числом или None.
    """
    if not value:
        return None
    v = str(value).strip()
    if not v:
        return None
    # "['1']" / '["1"]' / "[1]" -> "1"
    for ch in ["[", "]", "'", '"', " "]:
        v = v.replace(ch, "")
    # после чистки может остаться "1,2" - нам нужно одно значение
    if "," in v:
        v = v.split(",", 1)[0].strip()
    return v or None


def _parse_int_set_csv(value: str | None) -> set[int]:
    if not value:
        return set()
    out: set[int] = set()
    for part in str(value).split(","):
        part = part.strip()
        if not part:
            continue
        try:
            out.add(int(part))
        except ValueError:
            continue
    return out


def _serialize_int_set_csv(values: set[int]) -> str:
    return ",".join(str(x) for x in sorted(values))


def _safe_int(value, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


# -----------------------------
# Filters
# -----------------------------
class DepartmentFilter(admin.SimpleListFilter):
    title = "Подразделение"
    parameter_name = "dep"  # ✅ НЕ "department", чтобы не было конфликтов

    def lookups(self, request, model_admin):
        qs = (
            CurrentStock.objects.select_related("department")
            .values_list("department_id", "department__name")
            .distinct()
            .order_by("department__name")
        )
        return [(str(dep_id), dep_name) for dep_id, dep_name in qs if dep_id]

    def queryset(self, request, queryset):
        dep_id = _clean_single_int(self.value())
        if dep_id:
            return queryset.filter(department_id=dep_id)
        return queryset


class WarehouseMultiFilter(admin.SimpleListFilter):
    title = "Склад"
    parameter_name = "warehouses"  # CSV: "2,5,9"

    def lookups(self, request, model_admin):
        dep_id = _clean_single_int(request.GET.get("dep"))

        qs = (
            CurrentStock.objects.select_related("warehouse")
            .values_list("warehouse_id", "warehouse__name", "warehouse__department_id")
            .distinct()
        )

        items = []
        seen = set()

        for wh_id, wh_name, wh_dep_id in qs:
            if not wh_id:
                continue
            if dep_id and str(wh_dep_id) != str(dep_id):
                continue
            if wh_id in seen:
                continue
            seen.add(wh_id)
            items.append((str(wh_id), wh_name))

        items.sort(key=lambda x: (x[1] or "").lower())
        return items

    def queryset(self, request, queryset):
        selected = _parse_int_set_csv(self.value())
        if selected:
            return queryset.filter(warehouse_id__in=selected)
        return queryset

    def choices(self, changelist):
        selected = _parse_int_set_csv(self.value())

        # сохраняем все параметры (поиск, dep, и т.д.)
        base_params = changelist.get_filters_params()
        base_params.pop(self.parameter_name, None)

        # All (сброс складов, но dep остаётся)
        yield {
            "selected": not selected,
            "query_string": "?" + urlencode(base_params),
            "display": "All",
        }

        for lookup, title in self.lookup_choices:
            try:
                wh_id = int(lookup)
            except ValueError:
                continue

            is_selected = wh_id in selected
            new_selected = set(selected)

            if is_selected:
                new_selected.remove(wh_id)
            else:
                new_selected.add(wh_id)

            params = dict(base_params)
            if new_selected:
                params[self.parameter_name] = _serialize_int_set_csv(new_selected)

            yield {
                "selected": is_selected,
                "query_string": "?" + urlencode(params),
                "display": f"{'☑' if is_selected else '☐'} {title}",
            }


@admin.register(CurrentStock)
class CurrentStockAdmin(admin.ModelAdmin):
    change_list_template = "admin/stock/currentstock/change_list.html"

    list_display = (
        "row_no",
        "department_name",
        "warehouse_name",
        "nomenclature_no",
        "name",
        "uom",
        "qty_fmt",
        "price_fmt",
        "amount_fmt",
        "months_no_move_fmt",
        "date_in",
        "date_out",
        "date_in_ngmk",
        "loaded_at",
    )

    list_filter = (DepartmentFilter, WarehouseMultiFilter)
    search_fields = ("nomenclature_no", "bso", "name")
    list_per_page = 50

    _row_counter = 0

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.display(description="№")
    def row_no(self, obj):
        self._row_counter += 1
        return self._row_counter

    def changelist_view(self, request, extra_context=None):
        page = int(request.GET.get("p", 0) or 0)
        self._row_counter = page * self.list_per_page
        return super().changelist_view(request, extra_context=extra_context)

    @admin.display(description="Подразделение", ordering="department__name")
    def department_name(self, obj: CurrentStock):
        return obj.department.name if obj.department_id else ""

    @admin.display(description="Склад", ordering="warehouse__name")
    def warehouse_name(self, obj: CurrentStock):
        return obj.warehouse.name if obj.warehouse_id else ""

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import/", self.admin_site.admin_view(self.import_redirect), name="currentstock_import"),
            path("export/", self.admin_site.admin_view(self.export_excel), name="currentstock_export"),

            # ✅ Endpoint для Excel-grid (Tabulator) внутри текущего changelist
            path("data/", self.admin_site.admin_view(self.grid_data), name="currentstock_grid_data"),
        ]
        return custom_urls + urls

    def import_redirect(self, request):
        return redirect("/admin/stock/import-uat/")

    # -----------------------------
    # форматирование RU
    # -----------------------------
    def _fmt_ru(self, v, decimals: int) -> str:
        if v is None:
            return ""
        s = f"{v:,.{decimals}f}"
        s = s.replace(",", "X").replace(".", ",").replace("X", "\u202F")
        return format_html('<span style="white-space:nowrap">{}</span>', s)

    @admin.display(description="Кол-во", ordering="qty")
    def qty_fmt(self, obj: CurrentStock):
        return self._fmt_ru(obj.qty, 3)

    @admin.display(description="Цена", ordering="price")
    def price_fmt(self, obj: CurrentStock):
        return self._fmt_ru(obj.price, 2)

    @admin.display(description="Сумма", ordering="amount")
    def amount_fmt(self, obj: CurrentStock):
        return self._fmt_ru(obj.amount, 2)

    @admin.display(description="Мес. без движения", ordering="months_no_move")
    def months_no_move_fmt(self, obj: CurrentStock):
        v = obj.months_no_move
        if v is None:
            return ""
        return f"{v:.1f}".replace(".", ",")

    # -----------------------------
    # Excel Grid JSON (Tabulator)
    # -----------------------------
    def _request_without_tabulator_params(self, request):
        """
        Django admin ChangeList падает, если в GET есть неизвестные параметры (например page/size).
        Поэтому для построения ChangeList делаем копию request и чистим GET.
        """
        req = copy.copy(request)
        qd = request.GET.copy()

        # Tabulator params
        for k in ("page", "size", "sorters", "filters", "sortField", "sortDir"):
            qd.pop(k, None)

        # На всякий случай: если когда-то добавим другие grid-параметры
        # (можешь расширить список)
        req.GET = qd
        return req

    def grid_data(self, request):
        """
        Возвращает JSON для Tabulator.
        Важно: queryset берём через Django changelist, значит:
        - учитывается правый FILTER (dep, warehouses)
        - учитывается поиск (q)
        - любые будущие list_filter / search_fields тоже будут учитываться
        """

        # ✅ FIX: убираем page/size и прочие параметры Tabulator ДО ChangeList
        clean_request = self._request_without_tabulator_params(request)

        # 1) Базовый queryset как в админке (со всеми фильтрами/поиском)
        cl = self.get_changelist_instance(clean_request)
        qs = cl.get_queryset(clean_request)

        # 2) Tabulator pagination
        page = _safe_int(request.GET.get("page"), 1)
        size = _safe_int(request.GET.get("size"), 50)
        if page < 1:
            page = 1
        if size < 1:
            size = 50
        if size > 5000:
            size = 5000  # защита, чтобы случайно не грузить всё

        # 3) Sorting
        # Tabulator может прислать:
        # - sorters=[{"field":"qty","dir":"asc"}]
        # или:
        # - sortField=qty&sortDir=asc
        sorters_raw = request.GET.get("sorters")
        sort_field = request.GET.get("sortField")
        sort_dir = request.GET.get("sortDir")

        ordering = []
        if sorters_raw:
            try:
                sorters = json.loads(sorters_raw)
                if isinstance(sorters, list) and sorters:
                    for s in sorters:
                        field = (s.get("field") or "").strip()
                        direc = (s.get("dir") or "asc").strip().lower()
                        if not field:
                            continue
                        prefix = "-" if direc == "desc" else ""
                        ordering.append(prefix + field)
            except Exception:
                ordering = []

        if not ordering and sort_field:
            direc = (sort_dir or "asc").strip().lower()
            prefix = "-" if direc == "desc" else ""
            ordering = [prefix + sort_field]

        if ordering:
            try:
                qs = qs.order_by(*ordering)
            except Exception:
                pass

        # 4) Column filters
        filters_raw = request.GET.get("filters")
        if filters_raw:
            try:
                filters = json.loads(filters_raw)
                if isinstance(filters, list):
                    for f in filters:
                        field = (f.get("field") or "").strip()
                        ftype = (f.get("type") or "=").strip().lower()
                        value = f.get("value")

                        if not field:
                            continue
                        if value is None or value == "":
                            continue

                        lookup = field
                        negate = False

                        if ftype in ("=", "eq"):
                            lookup = field
                        elif ftype in ("!=", "neq", "ne"):
                            lookup = field
                            negate = True
                        elif ftype in ("like", "contains"):
                            lookup = f"{field}__icontains"
                        elif ftype in (">", "gt"):
                            lookup = f"{field}__gt"
                        elif ftype in ("<", "lt"):
                            lookup = f"{field}__lt"
                        elif ftype in (">=", "gte"):
                            lookup = f"{field}__gte"
                        elif ftype in ("<=", "lte"):
                            lookup = f"{field}__lte"
                        else:
                            lookup = field

                        try:
                            if negate:
                                qs = qs.exclude(**{lookup: value})
                            else:
                                qs = qs.filter(**{lookup: value})
                        except Exception:
                            pass
            except Exception:
                pass

        total = qs.count()
        start = (page - 1) * size
        end = start + size

        rows = []
        for obj in qs.select_related("department", "warehouse")[start:end]:
            rows.append({
                "id": obj.pk,
                "department": obj.department.name if obj.department_id else "",
                "warehouse": obj.warehouse.name if obj.warehouse_id else "",
                "nomenclature_no": obj.nomenclature_no or "",
                "name": obj.name or "",
                "uom": obj.uom or "",
                "qty": float(obj.qty) if obj.qty is not None else None,
                "price": float(obj.price) if obj.price is not None else None,
                "amount": float(obj.amount) if obj.amount is not None else None,
                "months_no_move": float(obj.months_no_move) if obj.months_no_move is not None else None,
                "date_in": obj.date_in.isoformat() if obj.date_in else None,
                "date_out": obj.date_out.isoformat() if obj.date_out else None,
                "date_in_ngmk": obj.date_in_ngmk.isoformat() if obj.date_in_ngmk else None,
                "loaded_at": obj.loaded_at.isoformat() if obj.loaded_at else None,
            })

        return JsonResponse({
            "data": rows,
            "last_page": (total // size) + (1 if total % size else 0),
            "total": total,
            "page": page,
            "size": size,
        }, json_dumps_params={"ensure_ascii": False})

    # -----------------------------
    # Export Excel
    # -----------------------------
    def export_excel(self, request):
        cl = self.get_changelist_instance(request)
        qs = cl.get_queryset(request)

        wb = Workbook()
        ws = wb.active
        ws.title = "CurrentStock"

        headers = [
            "Подразделение",
            "Склад",
            "Номенкл №",
            "Наименование",
            "Ед. изм.",
            "Кол-во",
            "Цена",
            "Сумма",
            "Мес. без движения",
            "Дата прихода на склад",
            "Дата расхода",
            "Дата прихода в НГМК",
        ]
        ws.append(headers)

        for obj in qs.iterator(chunk_size=2000):
            ws.append([
                obj.department.name if obj.department_id else "",
                obj.warehouse.name if obj.warehouse_id else "",
                obj.nomenclature_no or "",
                obj.name or "",
                obj.uom or "",
                float(obj.qty) if obj.qty is not None else None,
                float(obj.price) if obj.price is not None else None,
                float(obj.amount) if obj.amount is not None else None,
                float(obj.months_no_move) if obj.months_no_move is not None else None,
                obj.date_in,
                obj.date_out,
                obj.date_in_ngmk,
            ])

        # Форматы
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = "#,##0.000"
        for col in (7, 8):
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = "#,##0.00"
        for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
            for cell in row:
                cell.number_format = "0.0"
        for col in (10, 11, 12):
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = "DD.MM.YYYY"

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        filename = "currentstock_export.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


register_stock_import_page(admin.site)