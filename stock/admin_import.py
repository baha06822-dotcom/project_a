from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.shortcuts import render, redirect
from django.urls import path

from openpyxl import load_workbook

from warehouse.models import Department, Warehouse
from .models import CurrentStock
from .utils import extract_warehouse_number


class StockImportForm(forms.Form):
    department = forms.ModelChoiceField(
        label="Подразделение",
        queryset=Department.objects.filter(is_active=True).order_by("code"),
    )
    file = forms.FileField(label="Excel файл (.xlsx)")


def _safe_str(v) -> str:
    return ("" if v is None else str(v)).strip()


def _safe_decimal(v):
    """
    Возвращает Decimal или None. Подходит для DecimalField.
    Поддержка форматов:
      - 25326.000
      - 25326,000
      - 25 326,000  (с пробелами как разделителями тысяч)
      - "  1,23  "
    """
    try:
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None

        # убираем пробелы/неразрывные пробелы (часто бывают в Excel)
        s = s.replace("\xa0", " ").replace(" ", "")

        # иногда приходит "1,23" вместо "1.23"
        s = s.replace(",", ".")

        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _safe_months(v):
    """
    Месяцев без движения: поддержка '9,5' / '9.5' / '9' / 9 / 9.5
    Возвращает Decimal с 1 знаком или None.
    """
    try:
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None

        s = s.replace("\xa0", " ").replace(" ", "")
        s = s.replace(",", ".")

        d = Decimal(s)
        return d.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def _safe_date(v):
    """
    Приводит значение из Excel к date или None.
    Поддержка: None / "" / datetime / date / "DD.MM.YYYY" / "YYYY-MM-DD"
    """
    if v is None:
        return None

    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # DD.MM.YYYY
        try:
            return datetime.strptime(s, "%d.%m.%Y").date()
        except Exception:
            pass
        # YYYY-MM-DD
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    if isinstance(v, datetime):
        return v.date()

    if isinstance(v, date):
        return v

    return None


@transaction.atomic
def import_uat_stock(*, department: Department, file_obj) -> tuple[int, int]:
    """
    Импорт УАТ-формата.

    Интересующие колонки:
    A: Склад (например '84 3709')
    B: Номенкл №
    C: БСО
    D: Наименование
    E: Ед. изм.
    F: Количество
    G: Цена
    H: Сумма
    L: Дата прихода на склад
    M: Дата расхода
    N: Дата прихода в НГМК
    O: Месяцев без движения
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    CurrentStock.objects.filter(department=department).delete()

    to_create: list[CurrentStock] = []
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if row is None or len(row) < 15:
                skipped += 1
                continue

            # A=0,B=1,C=2,D=3,E=4,F=5,G=6,H=7,L=11,M=12,N=13,O=14
            wh_raw = _safe_str(row[0])
            nom_no = _safe_str(row[1])

            if not wh_raw or not nom_no:
                skipped += 1
                continue

            wh_number = extract_warehouse_number(wh_raw)
            if not wh_number:
                skipped += 1
                continue

            wh_code = f"{department.code}_{wh_number}"
            warehouse = Warehouse.objects.get(department=department, code=wh_code)

            obj = CurrentStock(
                department=department,
                warehouse=warehouse,
                nomenclature_no=nom_no,
                bso=_safe_str(row[2]),
                name=_safe_str(row[3])[:512],
                uom=_safe_str(row[4])[:32],
                qty=_safe_decimal(row[5]),
                price=_safe_decimal(row[6]),
                amount=_safe_decimal(row[7]),
                date_in=_safe_date(row[11]),
                date_out=_safe_date(row[12]),
                date_in_ngmk=_safe_date(row[13]),
                months_no_move=_safe_months(row[14]),
            )
            to_create.append(obj)
            imported += 1

        except Warehouse.DoesNotExist:
            skipped += 1
        except Exception:
            skipped += 1

    CurrentStock.objects.bulk_create(to_create, batch_size=2000)
    return imported, skipped


@transaction.atomic
def import_cmtb_stock(*, department: Department, file_obj) -> tuple[int, int]:
    """
    Импорт формата ЦМТБ CRU (dep.code = CRU_CMTB).

    Колонки:
    A: Подразделение (игнорируем)
    B: Склад
    C: Номенклатура
    E: Ед. изм
    F: Цена
    G: Кол-во
    H: Сумма
    I: Наименование
    J: Дата (приход)
    K: Дата (расход)
    L: Месяц без движ
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    CurrentStock.objects.filter(department=department).delete()

    to_create: list[CurrentStock] = []
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if row is None or len(row) < 12:
                skipped += 1
                continue

            # A=0,B=1,C=2,E=4,F=5,G=6,H=7,I=8,J=9,K=10,L=11
            wh_raw = _safe_str(row[1])  # B
            nom_no = _safe_str(row[2])  # C

            if not wh_raw or not nom_no:
                skipped += 1
                continue

            wh_number = extract_warehouse_number(wh_raw)
            if not wh_number:
                skipped += 1
                continue

            wh_code = f"{department.code}_{wh_number}"
            warehouse = Warehouse.objects.get(department=department, code=wh_code)

            obj = CurrentStock(
                department=department,
                warehouse=warehouse,
                nomenclature_no=nom_no,
                bso="",
                name=_safe_str(row[8])[:512],
                uom=_safe_str(row[4])[:32],
                price=_safe_decimal(row[5]),
                qty=_safe_decimal(row[6]),
                amount=_safe_decimal(row[7]),
                date_in=_safe_date(row[9]),
                date_out=_safe_date(row[10]),
                date_in_ngmk=None,
                months_no_move=_safe_months(row[11]),
            )
            to_create.append(obj)
            imported += 1

        except Warehouse.DoesNotExist:
            skipped += 1
        except Exception:
            skipped += 1

    CurrentStock.objects.bulk_create(to_create, batch_size=2000)
    return imported, skipped


def _find_navoi_warehouse(*, department: Department, wh_number: str) -> Warehouse | None:
    """
    NAVOI_CMTB:
    1) сначала пробуем точный код: CMTB_NAVOI_2801
    2) если не найден — ищем по окончанию: *_2801
    """
    # 1) точный ожидаемый шаблон в твоей БД
    exact_code = f"CMTB_NAVOI_{wh_number}"
    warehouse = Warehouse.objects.filter(department=department, code=exact_code).first()
    if warehouse:
        return warehouse

    # 2) запасной вариант — по окончанию кода
    return Warehouse.objects.filter(department=department, code__endswith=f"_{wh_number}").first()


@transaction.atomic
def import_navoi_cmtb_stock(*, department: Department, file_obj) -> tuple[int, int]:
    """
    Импорт формата ЦМТБ Навои (dep.code = NAVOI_CMTB).

    Нужные колонки:
    D = Склад (текст, внутри есть номер склада)
    E = Номенкл
    H = Наименование
    I = Ед. изм
    K = Кол-во
    L = Цена
    M = Сумма
    N = Дата
    O = Дата
    AC = Месяц без движения

    Индексы 0-based:
    D=3, E=4, H=7, I=8, K=10, L=11, M=12, N=13, O=14, AC=28
    """
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active

    CurrentStock.objects.filter(department=department).delete()

    to_create: list[CurrentStock] = []
    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if row is None or len(row) < 29:
                skipped += 1
                continue

            wh_raw = _safe_str(row[3])   # D
            nom_no = _safe_str(row[4])   # E

            if not wh_raw or not nom_no:
                skipped += 1
                continue

            wh_number = extract_warehouse_number(wh_raw)
            if not wh_number:
                skipped += 1
                continue

            warehouse = _find_navoi_warehouse(department=department, wh_number=wh_number)
            if warehouse is None:
                skipped += 1
                continue

            obj = CurrentStock(
                department=department,
                warehouse=warehouse,
                nomenclature_no=nom_no,
                bso="",
                name=_safe_str(row[7])[:512],   # H
                uom=_safe_str(row[8])[:32],     # I
                qty=_safe_decimal(row[10]),     # K
                price=_safe_decimal(row[11]),   # L
                amount=_safe_decimal(row[12]),  # M
                date_in=_safe_date(row[13]),    # N
                date_out=_safe_date(row[14]),   # O
                date_in_ngmk=None,
                months_no_move=_safe_months(row[28]),  # AC
            )
            to_create.append(obj)
            imported += 1

        except Exception:
            skipped += 1

    CurrentStock.objects.bulk_create(to_create, batch_size=2000)
    return imported, skipped


def register_stock_import_page(admin_site: admin.AdminSite):
    def import_view(request):
        if request.method == "POST":
            form = StockImportForm(request.POST, request.FILES)
            if form.is_valid():
                dep = form.cleaned_data["department"]
                f = form.cleaned_data["file"]

                if dep.code == "CRU_UAT":
                    imported, skipped = import_uat_stock(department=dep, file_obj=f)
                elif dep.code == "CRU_CMTB":
                    imported, skipped = import_cmtb_stock(department=dep, file_obj=f)
                elif dep.code == "NAVOI_CMTB":
                    imported, skipped = import_navoi_cmtb_stock(department=dep, file_obj=f)
                else:
                    messages.error(
                        request,
                        "Импорт пока поддерживает только подразделения CRU_UAT, CRU_CMTB и NAVOI_CMTB.",
                    )
                    return redirect(request.path)

                messages.success(
                    request,
                    f"Импорт завершён. Загружено: {imported}. Пропущено: {skipped}.",
                )
                return redirect("/admin/stock/currentstock/")
        else:
            form = StockImportForm()

        context = {
            **admin_site.each_context(request),
            "title": "Импорт остатков (Excel)",
            "form": form,
        }
        return render(request, "admin/stock_import.html", context)

    original_get_urls = admin_site.get_urls

    def get_urls():
        custom = [
            path("stock/import-uat/", admin_site.admin_view(import_view), name="stock_import_uat"),
        ]
        return custom + original_get_urls()

    admin_site.get_urls = get_urls